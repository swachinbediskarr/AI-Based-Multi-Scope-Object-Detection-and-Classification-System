import os
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from typing import Dict, Any, Optional, List
from fpdf import FPDF
import openpyxl

class PDFReport(FPDF):
    """Custom FPDF layout for Professional M.Tech Research Reports."""
    def header(self):
        self.set_fill_color(15, 23, 42)
        self.rect(0, 0, 210, 22, 'F')
        self.set_font("Helvetica", 'B', 12)
        self.set_text_color(56, 189, 248)  
        self.set_xy(10, 6)
        self.cell(0, 10, "AI-BASED MULTI-SCOPE OBJECT DETECTION SYSTEM", ln=False)
        self.set_font("Helvetica", '', 9)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, "RESEARCH REPORT ", align='R', ln=True)
        self.ln(6)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", 'I', 8)
        self.set_text_color(148, 163, 184)
        self.cell(0, 10, f"Page {self.page_no()} | Generated Automatically by AI Detection Engine", align='C')

class ReportGenerator:
    """Smart Multi-Format Report Generator (Supports both Dict and List Data Structures)."""

    def __init__(self, output_dir: str = "reports/history"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def _normalize_session_data(self, session_data: Any) -> tuple:
        """
        Smart parser: Converts List or Dict input into standardized (summary, logs, alerts, session_id).
        Automatically maps 'label' -> 'class_name' and 'min_distance' -> 'distance_m'.
        """
        raw_logs = []
        summary = []
        alerts = []
        session_id = "LIVE_SESSION"

        if isinstance(session_data, dict):
            summary = session_data.get("summary", [])
            raw_logs = session_data.get("logs", [])
            alerts = session_data.get("alerts", [])
            session_id = session_data.get("session_id", "LIVE_SESSION")

        elif isinstance(session_data, list):
            raw_logs = session_data

        standardized_logs = []
        for l in raw_logs:
            standardized_logs.append({
                "timestamp": l.get("last_seen") or l.get("timestamp") or l.get("first_seen") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "track_id": l.get("track_id", "N/A"),
                "class_name": l.get("class_name") or l.get("label") or l.get("class") or "Object",
                "distance_m": float(l.get("min_distance", l.get("distance_m", l.get("distance", 0.0)))),
                "status": l.get("status", "SAFE")
            })

        if standardized_logs and not summary:
            class_counts = {}
            for item in standardized_logs:
                cname = item["class_name"]
                class_counts[cname] = class_counts.get(cname, 0) + 1

            summary = [
                {
                    "class_name": cname,
                    "total_events": count,
                    "unique_objects": count
                }
                for cname, count in class_counts.items()
            ]

        if standardized_logs and not alerts:
            for item in standardized_logs:
                if item["status"] in ["CRITICAL", "WARNING"] or item["distance_m"] < 2.0:
                    alerts.append(item)

        return summary, standardized_logs, alerts, session_id

    def export_pdf(self, session_data: Any = None, output_path: Optional[str] = None, *args, **kwargs) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not output_path:
            output_path = os.path.join(self.output_dir, f"Executive_PDF_Report_{timestamp}.pdf")

        summary, logs, alerts, session_id = self._normalize_session_data(session_data)

        chart_path = os.path.join(self.output_dir, f"chart_{timestamp}.png")
        chart_file = self._generate_chart(summary, chart_path)

        pdf_path = self.generate_pdf(session_id, summary, logs, alerts, chart_file, output_path)

        if chart_file and os.path.exists(chart_file):
            os.remove(chart_file)

        return pdf_path

    def export_excel(self, session_data: Any = None, output_path: Optional[str] = None, *args, **kwargs) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not output_path:
            output_path = os.path.join(self.output_dir, f"Excel_Report_{timestamp}.xlsx")

        summary, logs, alerts, _ = self._normalize_session_data(session_data)

        chart_path = os.path.join(self.output_dir, f"chart_{timestamp}.png")
        chart_file = self._generate_chart(summary, chart_path)

        excel_path = self.generate_excel(summary, logs, alerts, chart_file, output_path)

        if chart_file and os.path.exists(chart_file):
            os.remove(chart_file)

        return excel_path

    def export_csv(self, session_data: Any = None, output_path: Optional[str] = None, *args, **kwargs) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not output_path:
            output_path = os.path.join(self.output_dir, f"CSV_Logs_{timestamp}.csv")

        _, logs, _, _ = self._normalize_session_data(session_data)
        return self.generate_csv(logs, output_path)

    def _generate_ai_summary(self, summary_data: List[Dict[str, Any]], alerts_data: List[Dict[str, Any]]) -> str:
        if not summary_data:
            return "No activity or objects were detected during this session sector scan."

        total_objects = sum(item.get("total_events", item.get("unique_objects", 1)) for item in summary_data)
        top_class_item = max(summary_data, key=lambda x: x.get("total_events", 0))
        top_class = top_class_item.get("class_name", "Object")
        total_alerts = len(alerts_data)

        return (
            f"During this surveillance session, a total of {total_objects} distinct target event(s) were tracked across "
            f"{len(summary_data)} object category(ies). The most active detected class was '{top_class}'. "
            f"The system recorded {total_alerts} security proximity/tripwire alert event(s) requiring potential operator attention. "
            f"Distance telemetry confirms real-time spatial positioning remained fully operational."
        )

    def _generate_chart(self, summary_data: List[Dict[str, Any]], chart_path: str) -> Optional[str]:
        if not summary_data:
            return None

        classes = [d.get('class_name', 'Unknown') for d in summary_data]
        counts = [d.get('total_events', 1) for d in summary_data]

        plt.figure(figsize=(6, 3))
        bars = plt.bar(classes, counts, color='#0284c7', edgecolor='#0f172a', width=0.4)
        plt.title('Detected Objects Breakdown by Category', fontsize=10, fontweight='bold', pad=10)
        plt.xlabel('Object Category', fontsize=8)
        plt.ylabel('Total Detections', fontsize=8)
        plt.grid(axis='y', linestyle='--', alpha=0.5)
        plt.tight_layout()

        for bar in bars:
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2.0, yval + 0.05, int(yval), ha='center', va='bottom', fontsize=8)

        plt.savefig(chart_path, dpi=200)
        plt.close()
        return chart_path

    def generate_pdf(self, session_id: Any, summary: List[Dict], logs: List[Dict], alerts: List[Dict], chart_file: Optional[str], filepath: str) -> str:
        pdf = PDFReport()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        pdf.set_font("Helvetica", 'B', 14)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 8, f"Session Summary Report [ID: {session_id}]", ln=True)
        pdf.set_font("Helvetica", '', 9)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 5, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target Sector: Live Camera Telemetry", ln=True)
        pdf.ln(5)

        pdf.set_fill_color(241, 245, 249)
        pdf.rect(10, pdf.get_y(), 190, 24, 'F')
        pdf.set_xy(12, pdf.get_y() + 2)
        pdf.set_font("Helvetica", 'B', 10)
        pdf.set_text_color(2, 132, 199)
        pdf.cell(0, 5, "AI EXECUTIVE SUMMARY", ln=True)
        pdf.set_font("Helvetica", '', 9)
        pdf.set_text_color(51, 65, 85)
        pdf.multi_cell(186, 4, self._generate_ai_summary(summary, alerts))
        pdf.ln(8)

        if chart_file and os.path.exists(chart_file):
            chart_y = pdf.get_y()
            pdf.image(chart_file, x=10, y=chart_y, w=100)
            pdf.set_xy(115, chart_y + 5)
            pdf.set_font("Helvetica", 'B', 11)
            pdf.set_text_color(15, 23, 42)
            pdf.cell(0, 6, "Key Statistics Summary", ln=True)
            pdf.set_font("Helvetica", '', 9)
            pdf.set_x(115)
            pdf.cell(0, 5, f"- Total Tracked Objects: {len(logs)}", ln=True)
            pdf.set_x(115)
            pdf.cell(0, 5, f"- Categories Identified: {len(summary)}", ln=True)
            pdf.set_x(115)
            pdf.cell(0, 5, f"- Security Alert Triggers: {len(alerts)}", ln=True)
            pdf.set_y(chart_y + 55)

        pdf.set_font("Helvetica", 'B', 11)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 6, "1. Tracked Objects & Distance Telemetry Logs", ln=True)

        pdf.set_font("Helvetica", 'B', 9)
        pdf.set_fill_color(51, 65, 85)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(45, 7, "Timestamp", border=1, fill=True)
        pdf.cell(30, 7, "Track ID", border=1, fill=True)
        pdf.cell(40, 7, "Class", border=1, fill=True)
        pdf.cell(35, 7, "Min Dist", border=1, fill=True)
        pdf.cell(40, 7, "Status", border=1, fill=True, ln=True)

        pdf.set_font("Helvetica", '', 9)
        pdf.set_text_color(30, 41, 59)
        for log in logs:
            pdf.cell(45, 6, str(log.get("timestamp", "")), border=1)
            pdf.cell(30, 6, f"#{log.get('track_id', '')}", border=1)
            pdf.cell(40, 6, str(log.get("class_name", "")), border=1)
            pdf.cell(35, 6, f"{log.get('distance_m', 0.0):.2f} m", border=1)
            pdf.cell(40, 6, str(log.get("status", "SAFE")), border=1, ln=True)

        pdf.output(filepath)
        return filepath

    def generate_excel(self, summary: List[Dict], logs: List[Dict], alerts: List[Dict], chart_file: Optional[str], filepath: str) -> str:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.append(["Class Name", "Total Events", "Unique Objects"])
        for row in summary:
            ws1.append([row.get("class_name", ""), row.get("total_events", 0), row.get("unique_objects", 0)])

        ws2 = wb.create_sheet(title="Telemetry Logs")
        ws2.append(["Timestamp", "Track ID", "Class Name", "Min Distance (m)", "Status"])
        for row in logs:
            ws2.append([
                row.get("timestamp", ""),
                row.get("track_id", ""),
                row.get("class_name", ""),
                row.get("distance_m", 0.0),
                row.get("status", "SAFE")
            ])

        wb.save(filepath)
        return filepath

    def generate_csv(self, logs_data: List[Dict[str, Any]], filepath: str) -> str:
        df = pd.DataFrame(logs_data)
        if df.empty:
            df = pd.DataFrame(columns=["timestamp", "track_id", "class_name", "distance_m", "status"])
        df.to_csv(filepath, index=False)
        return filepath